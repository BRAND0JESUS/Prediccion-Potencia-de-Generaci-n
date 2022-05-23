# -*- coding: utf-8 -*-
"""
Created on Thu Mar  5 17:40:38 2020

@author: Brando

"""

#%% IMPORTACION DE LIBRERIAS A UTILIZAR
# Importa modulo os (comandos referentes al sistema operativo)
import os 
# Importar modulo pandas para importacion/exportacion de archivos
import pandas as pd
# Importar modulo numpy para uso de matrices
import numpy as np
# Importar StandardScaler y DBSCAN para encontrar valores atipicos normalizacion de los datos
from datetime import datetime
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import DBSCAN


from openpyxl import load_workbook
import matplotlib.pyplot as plt

# import statsmodels.formula.api as sm

## import statsmodels.formula.api as sm

#
#path = os.getcwd()
## import  statsmodels.api as sm

#%% RUTINA PARA PODER CAMBIAR DE CARPETA AL PROGRAMA
def Analisis_Data ():
    # Ubicar PATH actual
    PATH = os.path.dirname(os.path.abspath(__file__))
    # Direcciona nobre de archivos al PATH
    os.chdir(PATH) 
    np.random.seed(7)

    #%% IMPORTACION DE DATOS

    def Import_Datos (Dat):
        # Imputar valores faltantes
        HistData = Dat.interpolate(method='linear', limit_direction='forward')
        Fil, Col = HistData.shape  # Extraer el numero de filas y columnas
        Titulos = HistData.columns     # Titulo de las columnas; (Titulos[0]) toma el dato en la primera posicion
        # Separacion de la base de datos inicial
        Demand = HistData.loc[:, (Titulos[-1])].values.reshape(-1,1) # Vector Demanda electrica
        Production = pd.DataFrame(HistData.loc[:,(Titulos[1]) : (Titulos[Col-2])]) # Matriz Produccion por bloque
        V_Production = Production.values # Matriz Produccion por bloque
        Time = HistData.loc[:, 'FECHA'].values.reshape(-1,1)  # Vector dato de tiempo 
        # Suma de produccion total de petroleo por dia
        Total_Production = V_Production.sum(axis=1).reshape(-1,1)
        return (Col, Titulos, Time, V_Production, Demand, Total_Production) 

    #%% DETECCION VALORES ATIPICOS
    def Val_Atip (matz1):
        Col, Titulos, Time, V_Production, Demand, Total_Production = matz1
        # Rutina para identificar valores atipicosHistorical_Data_Clustering
        Historical_Data = np.concatenate((Total_Production, Demand), axis=1) # Concatenacion de produccion y demanda para el ingreso a la funcion de agrupamiento
        Historical_Data_Norm = StandardScaler().fit_transform(Historical_Data) # Normalizacion de los datos para mejorar el agrupamiento
        Historical_Data_Clustering = DBSCAN(eps=0.5,min_samples=5).fit(Historical_Data_Norm) # Agrupamiento de los datos considerando un radio de 0.3 y numero minimo de elementos por grupo de 5
    
        #%% VISUALIZACION GRAFICA DE LOS DATOS EN ESTUDIO IDENTIFICANDO VALORES ATIPICOS
        plt.figure(figsize =(16, 8))
        plt.title('Datos Totales')
        plt.xlabel('Produccion BBL')
        plt.ylabel('Potencia de Generacion MW')
        core_samples_mask = np.zeros_like(Historical_Data_Clustering.labels_, dtype=bool)
        core_samples_mask[Historical_Data_Clustering.core_sample_indices_] = True
        labels = Historical_Data_Clustering.labels_ # Matriz que almacena a que grupo pertenece cada dato
        unique_labels = set(labels) # Ve los grupos que agrupaciones existen (0,1,2), -1 son valores atipicos
        colors = [plt.cm.Spectral(each)
                  for each in np.linspace(0, 1, len(unique_labels))]  # Crea una combinacion de colores para graficar dependiendo del grupo al que pertenece 
        for k, col in zip(unique_labels, colors): # zip iteracion paralela de grupo con color asignado
            if k == -1:
                col = [1, 1, 0, 1] # Color usado para los valores atipicos amarillo

            class_member_mask = (labels == k)
            
            xy = Historical_Data[class_member_mask & core_samples_mask]
            plt.plot(xy[:, 0], xy[:, 1], 'o', markerfacecolor=tuple(col),
                     markeredgecolor='k')
        
            xy = Historical_Data[class_member_mask & ~core_samples_mask]
            plt.plot(xy[:, 0], xy[:, 1], 'o', markerfacecolor=tuple(col),
                     markeredgecolor='k')
        plt.show()
        return (labels)

    #%% ELIMINACION DE VALORES ATIPICOS

    def Elim_Atip (matz1, labels, name_Act, name_Atip):
        Col, Titulos, Time, V_Production, Demand, Total_Production = matz1
        fil_Prod,col_Prod = V_Production.shape
        New_Outliers = pd.read_excel('Prod_Activos.xlsx', sheet_name=name_Act, header=0)
        New_Outliers['FECHA'] = New_Outliers['FECHA'].dt.strftime('%m/%d/%Y')
        for i in range(0,len(labels)):
            if labels[i] != -1:
                New_Outliers = New_Outliers.drop([i],axis=0)
        Clear_sheet = pd.DataFrame(np.full((200,col_Prod+1),None)) #Crea matriz con un numero especifico- 2 filas 2 columnas
        book = load_workbook('Prod_Activos.xlsx')
        writer = pd.ExcelWriter('Prod_Activos.xlsx', engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        Clear_sheet.to_excel(writer, sheet_name=name_Atip,  index = False) 
        New_Outliers.to_excel(writer, sheet_name=name_Atip,  index = False) 
        writer.save()
    
      
        ## Eliminacion de Valores atipicos de Base de Datos
        Hist_Atip = pd.read_excel('Prod_Activos.xlsx', sheet_name=name_Atip, header=0)
        # Imputar valores faltantes
        Hist_Atip = Hist_Atip.interpolate(method='linear', limit_direction='forward')

        Demand_Atip = Hist_Atip.loc[:, (Titulos[-1])].values.reshape(-1,1) # Vector Demanda electrica
        Production_Atip = pd.DataFrame(Hist_Atip.loc[:,(Titulos[1]) : (Titulos[Col-2])]) # Matriz Produccion por bloque
        V_Production_Atip = Production_Atip.values # Matriz Produccion por bloque
        Time_Atip = Hist_Atip.loc[:, 'FECHA'].values.reshape(-1,1)  # Vector dato de tiempo 

    # Eliminacion de atipicos de Produccion
        for m in range(len(V_Production_Atip)):
            for n in range(len(V_Production_Atip[m])):
                for i in range(len(V_Production)):
                    for j in range(len(V_Production[i])):
                        if name_Act == 'CUYABENO' or  name_Act == 'LIBERTADOR':
                            if V_Production_Atip[m][2]==V_Production[i][2]:
                                V_Production[i][j]=10101  # Remplaza los valores repetidos por 10101
                        else:
                            if V_Production_Atip[m][0]==V_Production[i][0]:
                                V_Production[i][j]=10101  # Remplaza los valores repetidos por 10101
                                

        rows_Prod = np.where(V_Production[:,0] != 10101)  # Toma las columnas que no poseen el valor 10101
        V_Production = V_Production[rows_Prod]

        # Eliminacion de atipicos de Potencia
        for p in range(len(Demand_Atip)):
            for q in range(len(Demand)):
                if Demand_Atip[p]==Demand[q]:
                    Demand[q]=101
        rows_Pot = np.where(Demand != 101)  # Toma las columnas que no poseen el valor 10101         
        Demand = (Demand[rows_Pot]).reshape(1, -1).T
        # Suma de produccion total de petroleo por dia
        Total_Production = (V_Production.sum(axis=1)).reshape(1, -1)
    
        return (V_Production, Demand)



    #%% PRINCIPAL

    Datos_Pred = pd.ExcelFile('Prediccion.xlsx')
    Name_Sheet_Predict = Datos_Pred.sheet_names
    Tot, AU, CY, EY, IN, ITT, LA, LI, OY, PA, SA, SH, AU_Pred, CY_Pred, EY_Pred, IN_Pred, ITT_Pred, LA_Pred, LI_Pred, OY_Pred, PA_Pred, SA_Pred, SH_Pred = Name_Sheet_Predict

    Datos = pd.ExcelFile('Prod_Activos.xlsx')
    Name_Sheet = Datos.sheet_names
    Act, AU, CY, EY, IN, ITT, LA, LI, OY, PA, SA, SH, AU_Atip, CY_Atip, EY_Atip, IN_Atip, ITT_Atip, LA_Atip, LI_Atip, OY_Atip, PA_Atip, SA_Atip, SH_Atip = Name_Sheet

    Datos_AUCA = Datos.parse('AUCA')
    matriz_AUCA_1 = Import_Datos (Datos_AUCA)  
    matriz_AUCA_2 = Val_Atip (matriz_AUCA_1)    
    matriz_AUCA_3 = Elim_Atip (matriz_AUCA_1,matriz_AUCA_2,AU,AU_Atip)
    a,b,Time,d, Demand_AU, Prod_Diaria_AU = matriz_AUCA_1

    Datos_CUYABENO = Datos.parse('CUYABENO')
    matriz_CUYABENO_1 = Import_Datos (Datos_CUYABENO)  
    matriz_CUYABENO_2 = Val_Atip (matriz_CUYABENO_1)    
    matriz_CUYABENO_3 = Elim_Atip (matriz_CUYABENO_1,matriz_CUYABENO_2,CY,CY_Atip)
    a,b,c,d, Demand_CY, Prod_Diaria_CY = matriz_CUYABENO_1

    Datos_EDEN_YUTURI = Datos.parse('EDEN YUTURI')
    matriz_EDEN_YUTURI_1 = Import_Datos (Datos_EDEN_YUTURI)  
    matriz_EDEN_YUTURI_2 = Val_Atip (matriz_EDEN_YUTURI_1)    
    matriz_EDEN_YUTURI_3 = Elim_Atip (matriz_EDEN_YUTURI_1,matriz_EDEN_YUTURI_2,EY,EY_Atip) 
    a,b,c,d, Demand_EY, Prod_Diaria_EY = matriz_EDEN_YUTURI_1

    Datos_INDILLANA = Datos.parse('INDILLANA')
    matriz_INDILLANA_1 = Import_Datos (Datos_INDILLANA)  
    matriz_INDILLANA_2 = Val_Atip (matriz_INDILLANA_1)    
    matriz_INDILLANA_3 = Elim_Atip (matriz_INDILLANA_1,matriz_INDILLANA_2,IN,IN_Atip) 
    a,b,c,d, Demand_IN, Prod_Diaria_IN = matriz_INDILLANA_1

    Datos_ITT = Datos.parse('ITT')
    matriz_ITT_1 = Import_Datos (Datos_ITT)  
    matriz_ITT_2 = Val_Atip (matriz_ITT_1)    
    matriz_ITT_3 = Elim_Atip (matriz_ITT_1,matriz_ITT_2,ITT,ITT_Atip) 
    a,b,c,d, Demand_ITT, Prod_Diaria_ITT = matriz_ITT_1

    Datos_LAGO_AGRIO = Datos.parse('LAGO AGRIO')
    matriz_LAGO_AGRIO_1 = Import_Datos (Datos_LAGO_AGRIO)  
    matriz_LAGO_AGRIO_2 = Val_Atip (matriz_LAGO_AGRIO_1)    
    matriz_LAGO_AGRIO_3 = Elim_Atip (matriz_LAGO_AGRIO_1,matriz_LAGO_AGRIO_2,LA,LA_Atip)
    a,b,c,d, Demand_LA, Prod_Diaria_LA = matriz_LAGO_AGRIO_1

    Datos_LIBERTADOR = Datos.parse('LIBERTADOR')
    matriz_LIBERTADOR_1 = Import_Datos (Datos_LIBERTADOR)  
    matriz_LIBERTADOR_2 = Val_Atip (matriz_LIBERTADOR_1)    
    matriz_LIBERTADOR_3 = Elim_Atip (matriz_LIBERTADOR_1,matriz_LIBERTADOR_2,LI,LI_Atip) 
    a,b,c,d, Demand_LI, Prod_Diaria_LI = matriz_LIBERTADOR_1

    Datos_OSO_YURALPA = Datos.parse('OSO YURALPA')
    matriz_OSO_YURALPA_1 = Import_Datos (Datos_OSO_YURALPA)  
    matriz_OSO_YURALPA_2 = Val_Atip (matriz_OSO_YURALPA_1)    
    matriz_OSO_YURALPA_3 = Elim_Atip (matriz_OSO_YURALPA_1,matriz_OSO_YURALPA_2,OY,OY_Atip) 
    a,b,c,d, Demand_OY, Prod_Diaria_OY = matriz_OSO_YURALPA_1

    Datos_PALO_AZUL = Datos.parse('PALO AZUL')
    matriz_PALO_AZUL_1 = Import_Datos (Datos_PALO_AZUL)  
    matriz_PALO_AZUL_2 = Val_Atip (matriz_PALO_AZUL_1)    
    matriz_PALO_AZUL_3 = Elim_Atip (matriz_PALO_AZUL_1,matriz_PALO_AZUL_2,PA,PA_Atip) 
    a,b,c,d, Demand_PA, Prod_Diaria_PA = matriz_PALO_AZUL_1

    Datos_SACHA = Datos.parse('SACHA')
    matriz_SACHA_1 = Import_Datos (Datos_SACHA)  
    matriz_SACHA_2 = Val_Atip (matriz_SACHA_1)    
    matriz_SACHA_3 = Elim_Atip (matriz_SACHA_1,matriz_SACHA_2,SA,SA_Atip) 
    a,b,c,d, Demand_SA, Prod_Diaria_SA = matriz_SACHA_1

    Datos_SHUSHUFINDI = Datos.parse('SHUSHUFINDI')
    matriz_SHUSHUFINDI_1 = Import_Datos (Datos_SHUSHUFINDI)  
    matriz_SHUSHUFINDI_2 = Val_Atip (matriz_SHUSHUFINDI_1)    
    matriz_SHUSHUFINDI_3 = Elim_Atip (matriz_SHUSHUFINDI_1,matriz_SHUSHUFINDI_2,SH,SH_Atip) 
    a,b,c,d, Demand_SH, Prod_Diaria_SH = matriz_SHUSHUFINDI_1
    
    Total_Prod_Hist = Prod_Diaria_AU+Prod_Diaria_CY+Prod_Diaria_EY+Prod_Diaria_IN+Prod_Diaria_ITT+Prod_Diaria_LA+Prod_Diaria_LI+Prod_Diaria_OY+Prod_Diaria_PA+Prod_Diaria_SA+Prod_Diaria_SH
    Total_Demand_Hist = Demand_AU+Demand_CY+Demand_EY+Demand_IN+Demand_ITT+Demand_LA+Demand_LI+Demand_OY+Demand_PA+Demand_SA+Demand_SH
    
    Date= Time.dt.strftime('%m/%d/%Y')
    
    
    #Exportacion de resultados de Total de Fluido y Potencia a Excel
    Data_Hist_Date = pd.DataFrame(Date)
    Data_Hist_Prod = pd.DataFrame(Total_Prod_Hist)
    Data_Hist_Demand = pd.DataFrame(Total_Demand_Hist)
    Clear_sheet = pd.DataFrame(np.full((500,3),None)) #Crea matriz con un numero especifico- 2 filas 2 columnas
    book = load_workbook('Prod_Activos.xlsx')
    writer = pd.ExcelWriter('Prod_Activos.xlsx', engine='openpyxl') 
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    writer.book = book
    Clear_sheet.to_excel(writer, sheet_name=Act,  index = False)
    Data_Hist_Date.to_excel(writer, sheet_name=Act,  index = False,startcol = 0) 
    Data_Hist_Prod.to_excel(writer, sheet_name=Act,  index = False,startcol = 1) 
    Data_Hist_Demand.to_excel(writer, sheet_name=Act,  index = False, startcol = 2) 
    writer.save()
  
    return (matriz_AUCA_3, matriz_CUYABENO_3, matriz_EDEN_YUTURI_3, matriz_INDILLANA_3, matriz_ITT_3, matriz_LAGO_AGRIO_3, matriz_LIBERTADOR_3, matriz_OSO_YURALPA_3, matriz_PALO_AZUL_3, matriz_SACHA_3, matriz_SHUSHUFINDI_3)
    