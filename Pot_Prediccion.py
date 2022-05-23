# -*- coding: utf-8 -*-
"""
Created on Sat Mar  7 21:56:58 2020

@author: Brando
"""
#%% IMPORTACION DE LIBRERIAS A UTILIZAR
# Importa modulo os (comandos referentes al sistema operativo)
import os 
# Importar modulo pandas para importacion/exportacion de archivos
import pandas as pd
# Importar modulo numpy para uso de matrices
import numpy as np
from datetime import datetime


from openpyxl import load_workbook

import matplotlib.pyplot as plt

from sklearn.externals import joblib


def Prediccion_Potencia ():
   
    #%% PREDICCION
    
    def Predict (name_Act,name_Predict):

        ForeData = pd.read_excel('Prediccion.xlsx', sheet_name=name_Act, header=0) 
        # Imputar valores faltantes
        ForeData = ForeData.interpolate(method='linear', limit_direction='forward')
        Fil, Col = ForeData.shape  # Extraer el numero de filas y columnas
        Titulos = ForeData.columns     # Titulo de las columnas; (Titulos[0]) toma el dato en la primera posicion
        # Separacion de la base de datos inicial
        Forecasting_Production = pd.DataFrame(ForeData.loc[:,(Titulos[1]) : (Titulos[Col-1])]).values # Matriz Produccion por bloque
        Time_Fore = ForeData.loc[:, 'FECHA'].values.reshape(-1,1)  # Vector dato de tiempo 
        # Suma de produccion total de petroleo por dia
        #Total_Forecasting_Production = Forecasting_Production.sum(axis=1)
        
        ## Normalizacion de los datos historicos
        Prod_Normalized_Fut = Forecasting_Production
        # Time = np.arange(len(Prod_Normalized_Fut)) 
        ## Prediccion de la demanda electrica
        joblib_model = joblib.load(name_Act)
        predictions_Norma = joblib_model.predict(Prod_Normalized_Fut)
        Dat_Pred = predictions_Norma[predictions_Norma.size-1]
        Pred = predictions_Norma
        
        # Grafica de potencia proyectada
        # plt.figure(figsize =(16, 8))
        # plt.title('PREDICCIONES DE LA DEMANDA DE ENERGIA ELECTRICA')
        # plt.plot(Time_Fore,Pred,c='red',label='Potencia electrica proyectada')
        # plt.legend(loc='upper left')
        # plt.show()
        
        ##Exportacion de resultados a Excel
        New_Prediction = pd.read_excel('Prediccion.xlsx', sheet_name=name_Act, header=0)
        New_Prediction['FECHA']= New_Prediction['FECHA'].dt.strftime('%m/%d/%Y')
        Data_Forecasting_Pro = pd.DataFrame(New_Prediction)
        Data_Forecasting_Dem = pd.DataFrame(Pred,columns=[(Titulos[-1])])
        fil_Pred,col_Pred = Data_Forecasting_Pro.shape
        Clear_sheet = pd.DataFrame(np.full((200,col_Pred+1),None)) #Crea matriz con un numero especifico- 2 filas 2 columnas
        book = load_workbook('Prediccion.xlsx')
        writer = pd.ExcelWriter('Prediccion.xlsx', engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        Clear_sheet.to_excel(writer, sheet_name=name_Predict,  index = False) 
        Data_Forecasting_Pro.to_excel(writer, sheet_name=name_Predict,  index = False) 
        Data_Forecasting_Dem.to_excel(writer, sheet_name=name_Predict,  index = False, startcol = col_Pred) 
        writer.save()
            
        return (Dat_Pred)

    #%% PRINCIPAL

    Datos_Pred = pd.ExcelFile('Prediccion.xlsx')
    Name_Sheet_Predict = Datos_Pred.sheet_names
    Tot, AU, CY, EY, IN, ITT, LA, LI, OY, PA, SA, SH, AU_Pred, CY_Pred, EY_Pred, IN_Pred, ITT_Pred, LA_Pred, LI_Pred, OY_Pred, PA_Pred, SA_Pred, SH_Pred = Name_Sheet_Predict
    
    matriz_AUCA_6 = Predict (AU,AU_Pred)

    matriz_CUYABENO_6 = Predict (CY,CY_Pred)

    matriz_EDEN_YUTURI_6 = Predict (EY,EY_Pred)
    
    matriz_INDILLANA_6 = Predict (IN,IN_Pred)
    
    matriz_ITT_6 = Predict (ITT,ITT_Pred)

    matriz_LAGO_AGRIO_6 = Predict (LA,LA_Pred)
    
    matriz_LIBERTADOR_6 = Predict (LI,LI_Pred)
    
    matriz_OSO_YURALPA_6 = Predict (OY,OY_Pred)
    
    matriz_PALO_AZUL_6 = Predict (PA,PA_Pred)
    
    matriz_SACHA_6 = Predict (SA,SA_Pred)
    
    matriz_SHUSHUFINDI_6 = Predict (SH,SH_Pred)
    
    #%% Exportacion de resultados a Excel de Tabla de resultados

    Datos_Pred = pd.ExcelFile('Prediccion.xlsx')
    Name_Sheet_Predict = Datos_Pred.sheet_names
    Tot, AU, CY, EY, IN, ITT, LA, LI, OY, PA, SA, SH, AU_Pred, CY_Pred, EY_Pred, IN_Pred, ITT_Pred, LA_Pred, LI_Pred, OY_Pred, PA_Pred, SA_Pred, SH_Pred = Name_Sheet_Predict

    Date_Prediction=Datos_Pred.parse(sheet_name=AU)
    Date_Prediction = Date_Prediction.interpolate(method='linear', limit_direction='forward')
    Date_V= Date_Prediction['FECHA'].dt.strftime('%m/%d/%Y')
    Date_Predict = pd.DataFrame([Date_V[Date_V.size-1]])

        
    Pot_Total = (matriz_AUCA_6+matriz_CUYABENO_6+matriz_EDEN_YUTURI_6+matriz_INDILLANA_6+matriz_ITT_6+matriz_LAGO_AGRIO_6+matriz_LIBERTADOR_6+matriz_OSO_YURALPA_6+matriz_PALO_AZUL_6+matriz_SACHA_6+matriz_SHUSHUFINDI_6)
    ACTIVOS = pd.DataFrame([AU, CY, EY, IN, ITT, LA, LI, OY, PA, SA, SH, 'RESULTADO'])
    Resultados = pd.DataFrame([matriz_AUCA_6, matriz_CUYABENO_6, matriz_EDEN_YUTURI_6, matriz_INDILLANA_6,
                               matriz_ITT_6, matriz_LAGO_AGRIO_6,matriz_LIBERTADOR_6, matriz_OSO_YURALPA_6,
                               matriz_PALO_AZUL_6, matriz_SACHA_6, matriz_SHUSHUFINDI_6, Pot_Total])

    Clear_sheet = pd.DataFrame(np.full((200,2),None)) #Crea matriz con un numero especifico- 2 filas 2 columnas
    book = load_workbook('Prediccion.xlsx')
    writer = pd.ExcelWriter('Prediccion.xlsx', engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    Clear_sheet.to_excel(writer, sheet_name='TOTAL',  index = False) 
    Date_Predict.to_excel(writer, sheet_name='TOTAL',  index = False, startcol = 0) 
    ACTIVOS.to_excel(writer, sheet_name='TOTAL',  index = False, startcol = 1) 
    Resultados.to_excel(writer, sheet_name='TOTAL',  index = False, startcol = 2) 
    writer.save()
 
    return ()